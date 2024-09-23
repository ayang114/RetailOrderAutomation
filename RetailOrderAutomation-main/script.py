import os
import pandas as pd
from flask import Flask, request, send_file, render_template
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)

# Directory to save the uploaded files
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create the upload directory if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part"
    
    file = request.files['file']

    if file.filename == '':
        return "No selected file"
    
    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Process the Excel file
        df = pd.read_excel(filepath)

        # Print the columns to check names
        print("All COLUMNS in the uploaded file:", df.columns.tolist())

        # Specify the columns you want to delete for the first layout
        columns_to_delete_first_layout = ['order-item-id', 'purchase-date', 'payments-date', 'reporting-date', 'promise-date', 'days-past-promise', 
                             'buyer-email', 'buyer-name', 'payment-method-details', 'cpf', 'buyer-phone-number', 'number-of-items', 
                             'product-name', 'quantity-shipped', 'quantity-to-ship', 'ship-service-level', 'ship-service-name', 
                             'ship-address-1', 'ship-address-2', 'ship-address-3', 'address-type', 'ship-city', 'ship-state', 
                             'ship-postal-code', 'ship-country', 'gift-wrap-type', 'gift-message-text', 'payment-method', 
                             'cod-collectible-amount', 'already-paid', 'payment-method-fee', 'customized-url', 'customized-page', 
                             'is-business-order', 'purchase-order-number', 'price-designation', 'is-prime', 'is-global-express', 
                             'is-premium-order', 'buyer-company-name', 'licensee-name', 'license-number', 'license-state', 
                             'license-expiration-date', 'is-replacement-order', 'is-exchange-order', 'original-order-id', 
                             'is-transparency', 'default-ship-from-address-name', 'default-ship-from-address-field-1', 
                             'default-ship-from-address-field-2', 'default-ship-from-address-field-3', 'default-ship-from-city', 
                             'default-ship-from-state', 'default-ship-from-country', 'default-ship-from-postal-code', 
                             'is-ispu-order', 'store-chain-store-id', 'is-buyer-requested-cancellation', 
                             'buyer-requested-cancel-reason', 'ioss-number', 'is-shipping-settings-automation-enabled', 
                             'ssa-carrier', 'ssa-ship-method', 'tax-collection-model', 'tax-collection-responsible-party', 
                             'verge-of-cancellation', 'verge-of-lateShipment', 'signature-confirmation-recommended']

        # Delete the specified columns for the first layout if they exist
        for column in columns_to_delete_first_layout:
            if column in df.columns:
                df.drop(columns=[column], inplace=True)
                print(f"Deleted column: {column}")
            else:
                print(f"Column '{column}' not found. No column deleted.")

        # Specify the desired order of columns for the first layout
        new_order_first_layout = [
            'order-id', 
            'recipient-name', 
            'sku', 
            'quantity-purchased',  # This will be changed to 'QTY' below
        ]

        # Reorder the DataFrame for the first layout
        df_first_layout = df[new_order_first_layout]
        
        # Change 'quantity-purchased' to 'QTY'
        df_first_layout.rename(columns={'quantity-purchased': 'Qty'}, inplace=True)

        print(f"Re-Arranged Column Ordering for First Layout")

        # Get the current date
        today = datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD

        # Create the output file path with today's date
        output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{today} - Amazon Order Report.xlsx')
        
        # Save the first layout to the output file
        df_first_layout.to_excel(output_filepath, index=False, sheet_name='Layout 1')

        # Load the workbook to create the second layout
        wb = load_workbook(output_filepath)
        
        # Create the first sheet
        ws1 = wb['Layout 1']
        
        # Define the border style for Layout 1
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply borders to header cells in Layout 1
        for c in range(1, df_first_layout.shape[1] + 1):
            cell = ws1.cell(row=1, column=c)
            cell.border = thin_border  # Add border to header
            cell.alignment = Alignment(horizontal='center')  # Center align header

        # Set font style for the header row
        red_bold_font = Font(color="FF0000", bold=True)

        # Apply font style to header cells
        for c in range(1, df_first_layout.shape[1] + 1):
            cell = ws1.cell(row=1, column=c)
            cell.font = red_bold_font  # Red bold font for the header

        # Apply borders and formatting to data cells in Layout 1 and center everything
        for r_idx in range(2, df_first_layout.shape[0] + 2):  # Starting from row 2
            for c_idx in range(1, df_first_layout.shape[1] + 1):
                cell = ws1.cell(row=r_idx, column=c_idx)
                cell.border = thin_border  # Add border to data cells
                cell.alignment = Alignment(horizontal='center')  # Center align data

        # Set the width of columns in Layout 1
        column_widths_layout_1 = {
            1: 25,  # Width for 'order-id'
            2: 25,  # Width for 'recipient-name'
            3: 25,  # Width for 'sku'
            4: 10   # Width for 'QTY'
        }

        for col, width in column_widths_layout_1.items():
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

        # Calculate the total of the 'Qty' column
        total_qty = df_first_layout['Qty'].sum()

        # Determine the next available row for total
        next_row = df_first_layout.shape[0] + 2  # +2 to account for header and data rows

        # Write the total in the appropriate cells
        ws1.cell(row=next_row, column=3, value="Total")  # Column C
        ws1.cell(row=next_row, column=4, value=total_qty)  # Column D

        # Format the "Total" row (e.g., bold font, border)
        total_font = Font(bold=True)
        for c in range(3, 5):  # Columns C and D
            cell = ws1.cell(row=next_row, column=c)
            cell.font = total_font
            cell.border = thin_border  # Add border to the total cells
            cell.alignment = Alignment(horizontal='center')  # Center align total cells

        ws2 = wb.create_sheet(title='Layout 2')
        # Create the second sheet

        # Process the DataFrame for the second layout
        df_second_layout = df.copy()

        # Delete specified columns for the second layout
        df_second_layout.drop(columns=['order-id', 'recipient-name'], inplace=True)

        # Split the 'sku' into separate columns, filling with None if there are fewer segments
        sku_split = df_second_layout['sku'].str.split('-', expand=True)

        # Assign to respective columns with safety checks
        df_second_layout['Style'] = sku_split[0]  # First part
        df_second_layout['Color'] = sku_split[1]  # Second part
        df_second_layout['Size'] = sku_split[2]    # Third part
        df_second_layout['Other'] = sku_split[3]   # Fourth part (after the third hyphen)

        # Retain the required columns and remove the original 'sku' column
        df_second_layout = df_second_layout[['Style', 'Color', 'Size', 'quantity-purchased', 'Other']]

        # Rename 'quantity-purchased' to 'QTY'
        df_second_layout.rename(columns={'quantity-purchased': 'Qty'}, inplace=True)

        # Set the 'Other' column values to blank (for rows where it previously had 'BD')
        df_second_layout['Other'] = ''

        # Rename the last column to "Packs from #1 or No Inv."
        df_second_layout.rename(columns={'Other': 'Packs from #1 or No Inv.'}, inplace=True)

        # Sort the DataFrame first by 'Style', then by 'Color', and then by 'Size' within the same 'Style' and 'Color'
        df_second_layout = df_second_layout.sort_values(by=['Style', 'Color', 'Size'], ascending=[True, True, True]).reset_index(drop=True)

        # Create a new list to store the rows for the new sheet
        new_rows = []
        previous_style = None

        # Iterate through the sorted DataFrame and add rows
        for r_idx, row in df_second_layout.iterrows():
            current_style = row['Style']
            
            # Check if the current style is different from the previous one
            if previous_style is not None and current_style != previous_style:
                # Add a blank row to create a newline
                new_rows.append(['', '', '', '', ''])  # Empty row with same number of columns
            
            # Append the current row to the new_rows list
            new_rows.append(row.tolist())
            
            # Update the previous_style for the next iteration
            previous_style = current_style

        # Write the new rows to the second sheet
        for r_idx, row in enumerate(new_rows):
            for c_idx, value in enumerate(row):
                ws2.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        # Set headers with bold red font size 14 for the second layout
        header_font = Font(color="FF0000", bold=True, size=14)  # Red color, bold, size 14
        nonheader_font = Font(color="000000", bold=False, size=14)  # Non-header font style

        for c in range(1, df_second_layout.shape[1] + 1):
            cell = ws2.cell(row=1, column=c, value=df_second_layout.columns[c-1])
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')  # Center align header
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))  # Add border to header cells

        # Define the border style for Layout 2
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply borders and center alignment for data cells in Layout 2
        for r_idx in range(2, len(new_rows) + 2):  # Starting from row 2
            for c_idx in range(1, df_second_layout.shape[1] + 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.font = nonheader_font
                cell.alignment = Alignment(horizontal='center')  # Center align data
                cell.border = thin_border  # Add border to data cells

        # Set the width of columns in Layout 2
        column_widths_layout_2 = {
            1: 20,  # Width for 'Style'
            2: 20,  # Width for 'Color'
            3: 15,  # Width for 'Size'
            4: 10,  # Width for 'QTY'
            5: 30   # Width for 'Packs from #1 or No Inv.'
        }

        for col, width in column_widths_layout_2.items():
            ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width

        # Calculate the total for the 'Qty' column
        total_qty = df_second_layout['Qty'].sum()

        # Append the "Total" label and the total value
        total_row_index = len(new_rows) + 2  # Adjust for header and data rows
        ws2.cell(row=total_row_index, column=3, value="Total")  # "Total" in the third column
        ws2.cell(row=total_row_index, column=4, value=total_qty)  # Total quantity in the fourth column

        # Center align the "Total" cell
        ws2.cell(row=total_row_index, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=total_row_index, column=4).alignment = Alignment(horizontal='center')

        # Save the changes to the workbook
        wb.save(output_filepath)


        return send_file(output_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)