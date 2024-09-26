import os
import pandas as pd
from flask import Flask, request, send_file, render_template
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)

# Directory to save the uploaded files
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create the upload directory if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part"
    
    file = request.files['file']

    if file.filename == '':
        return "No selected file"
    
    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Process the Excel file
        df = pd.read_excel(filepath)

        # Print the columns to check names
        print("All COLUMNS in the uploaded file:", df.columns.tolist())

        # Specify the columns you want to delete for the first layout
        columns_to_delete_first_layout = ['order-item-id', 'purchase-date', 'payments-date', 'reporting-date', 'promise-date', 'days-past-promise', 
                             'buyer-email', 'buyer-name', 'payment-method-details', 'cpf', 'buyer-phone-number', 'number-of-items', 
                             'product-name', 'quantity-shipped', 'quantity-to-ship', 'ship-service-level', 'ship-service-name', 
                             'ship-address-1', 'ship-address-2', 'ship-address-3', 'address-type', 'ship-city', 'ship-state', 
                             'ship-postal-code', 'ship-country', 'gift-wrap-type', 'gift-message-text', 'payment-method', 
                             'cod-collectible-amount', 'already-paid', 'payment-method-fee', 'customized-url', 'customized-page', 
                             'is-business-order', 'purchase-order-number', 'price-designation', 'is-prime', 'is-global-express', 
                             'is-premium-order', 'buyer-company-name', 'licensee-name', 'license-number', 'license-state', 
                             'license-expiration-date', 'is-replacement-order', 'is-exchange-order', 'original-order-id', 
                             'is-transparency', 'default-ship-from-address-name', 'default-ship-from-address-field-1', 
                             'default-ship-from-address-field-2', 'default-ship-from-address-field-3', 'default-ship-from-city', 
                             'default-ship-from-state', 'default-ship-from-country', 'default-ship-from-postal-code', 
                             'is-ispu-order', 'store-chain-store-id', 'is-buyer-requested-cancellation', 
                             'buyer-requested-cancel-reason', 'ioss-number', 'is-shipping-settings-automation-enabled', 
                             'ssa-carrier', 'ssa-ship-method', 'tax-collection-model', 'tax-collection-responsible-party', 
                             'verge-of-cancellation', 'verge-of-lateShipment', 'signature-confirmation-recommended']

        # Delete the specified columns for the first layout if they exist
        for column in columns_to_delete_first_layout:
            if column in df.columns:
                df.drop(columns=[column], inplace=True)
                print(f"Deleted column: {column}")
            else:
                print(f"Column '{column}' not found. No column deleted.")

        # Specify the desired order of columns for the first layout
        new_order_first_layout = [
            'order-id', 
            'recipient-name', 
            'sku', 
            'quantity-purchased',  # This will be changed to 'QTY' below
        ]

        # Reorder the DataFrame for the first layout
        df_first_layout = df[new_order_first_layout]
        
        # Change 'quantity-purchased' to 'QTY'
        df_first_layout.rename(columns={'quantity-purchased': 'Qty'}, inplace=True)

        print(f"Re-Arranged Column Ordering for First Layout")

        # Get the current date
        today = datetime.now().strftime("%Y%m%d")  # Format: YYYY-MM-DD

        # Create the output file path with today's date
        output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{today} - Amazon Order Report.xlsx')
        
        # Save the first layout to the output file
        df_first_layout.to_excel(output_filepath, index=False, sheet_name=today)

        # Load the workbook to create the second layout
        wb = load_workbook(output_filepath)
        
        # Create the first sheet
        ws1 = wb[today]
        
        # Define the border style for Layout 1
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply font style to header cells with size 14
        for c in range(1, df_first_layout.shape[1] + 1):
            cell = ws1.cell(row=1, column=c)
            cell.font = Font(color="FF0000", bold=True, size=14)  # Red bold font, size 14 for the header
            cell.border = thin_border  # Add border to header
            cell.alignment = Alignment(horizontal='center')  # Center align header

        # Apply borders, formatting, and font size 14 to data cells in Layout 1 and center everything
        for r_idx in range(2, df_first_layout.shape[0] + 2):  # Starting from row 2
            for c_idx in range(1, df_first_layout.shape[1] + 1):
                cell = ws1.cell(row=r_idx, column=c_idx)
                cell.font = Font(size=14)  # Font size 14 for the data cells
                cell.border = thin_border  # Add border to data cells
                cell.alignment = Alignment(horizontal='center')  # Center align data

        # Set the width of columns in Layout 1
        column_widths_layout_1 = {
            1: 30,  # Width for 'order-id'
            2: 30,  # Width for 'recipient-name'
            3: 30,  # Width for 'sku'
            4: 10   # Width for 'QTY'
        }

        for col, width in column_widths_layout_1.items():
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

        # Calculate the total of the 'Qty' column
        total_qty = df_first_layout['Qty'].sum()

        # Determine the next available row for total
        next_row = df_first_layout.shape[0] + 2  # +2 to account for header and data rows

        # Write the total in the appropriate cells
        ws1.cell(row=next_row, column=3, value="Total")  # Column C
        ws1.cell(row=next_row, column=4, value=total_qty)  # Column D

        # Format the "Total" row (e.g., bold font, border)
        total_font = Font(bold=True, size=14)
        for c in range(3, 5):  # Columns C and D
            cell = ws1.cell(row=next_row, column=c)
            cell.font = total_font  # Bold font, size 14 for the total row
            cell.border = thin_border  # Add border to the total cells
            cell.alignment = Alignment(horizontal='center')  # Center align total cells

        # After creating the second layout (ws2)
        ws2 = wb.create_sheet(title=f'{today}-')

        # Process the DataFrame for the second layout
        df_second_layout = df.copy()

        # Delete specified columns for the second layout
        df_second_layout.drop(columns=['order-id', 'recipient-name'], inplace=True)

        # Split the 'sku' into separate columns, filling with None if there are fewer segments
        sku_split = df_second_layout['sku'].str.split('-', expand=True)

        # Assign to respective columns with safety checks
        df_second_layout['Style'] = sku_split[0]  # First part
        df_second_layout['Color'] = sku_split[1]  # Second part
        df_second_layout['Size'] = sku_split[2]    # Third part
        df_second_layout['Other'] = sku_split[3]   # Fourth part (after the third hyphen)

        # Retain the required columns and remove the original 'sku' column
        df_second_layout = df_second_layout[['Style', 'Color', 'Size', 'quantity-purchased', 'Other']]

        # Rename 'quantity-purchased' to 'Qty'
        df_second_layout.rename(columns={'quantity-purchased': 'Qty'}, inplace=True)

        # Set the 'Other' column values to blank (for rows where it previously had 'BD')
        df_second_layout['Other'] = ''

        # Rename the last column to "Packs from #1 or No Inv."
        df_second_layout.rename(columns={'Other': 'Packs from #1 or No Inv.'}, inplace=True)

        # Sort the DataFrame first by 'Style', then by 'Color', and then by 'Size' within the same 'Style' and 'Color'
        df_second_layout = df_second_layout.sort_values(by=['Style', 'Color', 'Size'], ascending=[True, True, True]).reset_index(drop=True)

        # Create a new list to store the rows for the new sheet
        new_rows = []
        previous_style = None

        # Iterate through the sorted DataFrame and add rows
        for r_idx, row in df_second_layout.iterrows():
            current_style = row['Style']
            
            # Check if the current style is different from the previous one
            if previous_style is not None and current_style != previous_style:
                # Add a blank row to create a newline
                new_rows.append(['', '', '', '', ''])  # Empty row with same number of columns
            
            # Append the current row to the new_rows list
            new_rows.append(row.tolist())
            
            # Update the previous_style for the next iteration
            previous_style = current_style

        # Write the new rows to the second sheet
        for r_idx, row in enumerate(new_rows):
            for c_idx, value in enumerate(row):
                ws2.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        # Set headers with bold red font size 14 for the second layout
        header_font = Font(color="FF0000", bold=True, size=14)  # Red color, bold, size 14
        nonheader_font = Font(color="000000", bold=False, size=14)  # Non-header font style

        for c in range(1, df_second_layout.shape[1] + 1):
            cell = ws2.cell(row=1, column=c, value=df_second_layout.columns[c-1])
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')  # Center align header
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))  # Add border to header cells

        # Define the border style for Layout 2
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply borders and center alignment for data cells in Layout 2
        for r_idx in range(2, len(new_rows) + 2):  # Starting from row 2
            for c_idx in range(1, df_second_layout.shape[1] + 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.font = nonheader_font
                cell.alignment = Alignment(horizontal='center')  # Center align data
                cell.border = thin_border  # Add border to data cells

        # Set the width of columns in Layout 2
        column_widths_layout_2 = {
            1: 20,  # Width for 'Style'
            2: 20,  # Width for 'Color'
            3: 15,  # Width for 'Size'
            4: 10,   # Width for 'QTY',
            5: 45   # Width for 'Packs from #1 or No Inv.'            
        }

        for col, width in column_widths_layout_2.items():
            ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width

        # Calculate the total for the 'Qty' column
        total_qty = df_second_layout['Qty'].sum()

        # Append the "Total" label and the total value
        total_row_index = len(new_rows) + 2  # Adjust for header and data rows
        ws2.cell(row=total_row_index, column=3, value="Total")  # "Total" in the third column
        ws2.cell(row=total_row_index, column=4, value=total_qty)  # Total quantity in the fourth column

        # Format the "Total" row (bold, size 14, and borders)
        total_font = Font(bold=True, size=14)
        for col in [3, 4]:  # Columns for "Total" and the count
            cell = ws2.cell(row=total_row_index, column=col)
            cell.font = total_font  # Bold font, size 14
            cell.alignment = Alignment(horizontal='center')  # Center align the total cells
            cell.border = thin_border  # Add border to the "Total" and quantity cells

        # Center align the "Total" cell
        ws2.cell(row=total_row_index, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=total_row_index, column=4).alignment = Alignment(horizontal='center')

        # Create the third sheet
        ws3 = wb.create_sheet(title='Fx5Reformatted')

        # Process the DataFrame for the third layout
        df_third_layout = df_second_layout.copy()  # Start from the second layout DataFrame

        # Remove the 5th column ('Packs from #1 or No Inv.') to create the third layout
        df_third_layout.drop(columns=['Packs from #1 or No Inv.'], inplace=True)

        # Update the 'Style' column based on the conditions
        def update_style(row):
            style, size = row['Style'], row['Size']
            
            # Dictionary mapping original styles to their new values
            style_mapping = {
                'CB0536': 'CB0536',
                'CO078': 'CO078M',
                'CO078LEO': 'CO078Y/LEO',
                'CO079': 'CO079M',
                'CO129': 'CO129Y',
                'CO129LEO': 'CO129Y/LEO',
                'HB2122': 'HB2122',
                'HB2137': 'HB2137',
                'HB2137PL': 'HB2137PL222',
                'HB3134': 'HB3134',
                'HK8072': 'HK8072',
                'HK8246': 'HK8246',
                'HK8266': 'HK8266',
                'HK8266': 'HK8266',
                'KC003': 'KC003',
                'KC009': 'KC009',
                'MK0179': 'MK0179',
                'MK3104': 'MK3104',
                'MK3279': 'MK3279',
                'MK3392': 'MK3392',
                'MK3466': 'MK3466',
                'MK3466': 'MK3466',
                'MK3467': 'MK3467Y',
                'MK3506': 'MK3506',
                'MK3514': 'MK3514Y',
                'MK3515': 'MK3515',
                'MK3554': 'MK3554',
                'MK3558': 'MK3558Y',
                'MK3595': 'MK3595',
                'MK3637': 'MK3637Y',
                'MK3636': 'MK3636Y',
                'MK3659': 'MK3659',
                'MK3664': 'MK3664Y',
                'MK3664LEO': 'MK3664Y/LEO',
                'MK3673': 'MK3673',
                'MK3675': 'MK3675',
                'MK5178': 'MK5178',
                'MK5500': 'MK5500',
                'MK5502': 'MK5502',
                'MK8015': 'MK8015',
                'MK8080': 'MK8080',
                'MK8144': 'MK8144',
                'MK8213': 'MK8213',
                'MK8236': 'MK8236',
                'MK8558': 'MK8558Y',
                'MK5501': 'MK5501',
                'MK3664EMBO': 'MK3664EMBO',
                'MK32004CAT': 'MK32004CAT',
                'MK8281': 'MK8281',
                'MK3399': 'MK3399',
                'MK8143': 'MK8143',
                'MK8268': 'MK8268',
                'MK3349': 'MK3349',

            }
            # Update styles based on specific conditions
            if style in ['CO129', 'CO078', 'CO079', 'HB2137', 'MK3514', 'MK3636', 'MK3558', 'MK3467', 'MK8558', 'MK3514KID', 'MK5178KID'] and size in ['1X', '2X', '3X', '4X']:
                if style == 'CO129':
                    return 'CO129PL222'
                elif style == 'CO078':
                    return 'CO078PL'
                elif style == 'CO079':
                    return 'CO079PL'
                elif style == 'HB2137':
                    return 'HB2137PL222'
                elif style == 'MK3514':
                    return 'MK3514PL'
                elif style == 'MK3636':
                    return 'MK3636Y'
                elif style == 'MK3558':
                    return 'MK3558Y'
                elif style == 'MK3467':
                    return 'MK3467Y'
                elif style == 'MK8558':
                    return 'MK8558Y'
                elif style == 'MK3514KID':
                    return 'MK3514KID'
                elif style == 'MK5178KID':
                    return 'MK5178KID'
            
            # Return the mapped style or the original style if no conditions are met
            return style_mapping.get(style, style)
        
            # Example mapping for color changes
            color_mapping = {
                'OPL': 'OPAL',
                'AQA': 'AQUA'
            }

            # Function to modify data for the third layout
            def process_third_layout(data):
                modified_data = []
                
                for row in data:
                    # Check if the color is in the mapping dictionary
                    if row['color'] in color_mapping:
                        # Update the color based on the mapping
                        row['color'] = color_mapping[row['color']]
                    modified_data.append(row)
                
                return modified_data


            # Update based on the size for the specified styles
            # if style in style_mapping:
            #     return style_mapping[style]

            # Default return if no conditions are met
            return style

         # Apply the style update function
        df_third_layout['Style'] = df_third_layout.apply(update_style, axis=1)





        # Write the updated third layout to the new sheet
        for r_idx, row in df_third_layout.iterrows():
            for c_idx, value in enumerate(row):
                ws3.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        # Set headers with bold red font size 14 for the third layout
        for c in range(1, df_third_layout.shape[1] + 1):
            cell = ws3.cell(row=1, column=c, value=df_third_layout.columns[c-1])
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply borders and formatting to data cells in Layout 3
        for r_idx in range(2, df_third_layout.shape[0] + 2):  # Starting from row 2
            for c_idx in range(1, df_third_layout.shape[1] + 1):
                cell = ws3.cell(row=r_idx, column=c_idx)
                cell.font = nonheader_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Set the width of columns in Layout 3
        column_widths_layout_3 = {
            1: 20,  # Width for 'Style'
            2: 15,  # Width for 'Color'
            3: 15,  # Width for 'Size'
            4: 10   # Width for 'Qty'
        }

        for col, width in column_widths_layout_3.items():
            ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = width

        # Save the workbook after all modifications
        wb.save(output_filepath)

        # Send the modified file back to the user
        return send_file(output_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)